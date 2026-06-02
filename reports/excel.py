from datetime import datetime, timezone
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import REPORTS_DIR, STARTING_BALANCE_EUR
from database import db
from trading.portfolio import get_portfolio_summary


_GREEN = "FF198754"
_RED = "FFDC3545"
_BLUE = "FF0D6EFD"
_HEADER_BG = "FF212529"
_ALT_ROW = "FFF8F9FA"
_WHITE = "FFFFFFFF"


def _header_style(ws, row: int, columns: list[str]) -> None:
    fill = PatternFill("solid", fgColor=_HEADER_BG)
    font = Font(bold=True, color=_WHITE, size=11)
    for col_idx, _ in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="FF495057")
        cell.border = Border(bottom=thin)


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _pnl_color(ws, row: int, col: int, value: float) -> None:
    cell = ws.cell(row=row, column=col)
    if value > 0:
        cell.font = Font(color=_GREEN, bold=True)
    elif value < 0:
        cell.font = Font(color=_RED, bold=True)


def generate_report() -> Path:
    summary = get_portfolio_summary()
    transactions = db.get_transactions(limit=500)
    research = db.get_all_research()

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    path = REPORTS_DIR / f"trading_report_{ts}.xlsx"

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # ── Sheet 1: Overview ────────────────────────────────────────────────
        overview_data = {
            "Metric": [
                "Starting Balance", "Cash Available", "Invested Value",
                "Portfolio Market Value", "Total Portfolio Value",
                "Unrealised P&L", "Unrealised P&L %",
                "Overall Return %", "Number of Positions",
                "Report Generated",
            ],
            "Value": [
                f"€{STARTING_BALANCE_EUR:.2f}",
                f"€{summary.cash_eur:.2f}",
                f"€{summary.total_invested_eur:.2f}",
                f"€{summary.total_market_value_eur:.2f}",
                f"€{summary.total_portfolio_eur:.2f}",
                f"{'+'if summary.total_pnl_eur>=0 else ''}€{summary.total_pnl_eur:.2f}",
                f"{summary.total_pnl_pct:+.2f}%",
                f"{summary.overall_return_pct:+.2f}%",
                str(len(summary.positions)),
                datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
            ],
        }
        pd.DataFrame(overview_data).to_excel(writer, sheet_name="Overview", index=False)

        # ── Sheet 2: Positions ───────────────────────────────────────────────
        if summary.positions:
            pos_rows = [
                {
                    "Symbol": p.symbol,
                    "Name": p.name,
                    "Quantity": round(p.quantity, 4),
                    "Avg Cost (€)": round(p.avg_cost_eur, 4),
                    "Current Price (€)": round(p.current_price_eur, 4),
                    "Market Value (€)": round(p.market_value_eur, 2),
                    "Cost Basis (€)": round(p.cost_basis_eur, 2),
                    "Unrealised P&L (€)": round(p.unrealised_pnl_eur, 2),
                    "Unrealised P&L %": f"{p.unrealised_pnl_pct:+.2f}%",
                    "Today %": f"{p.change_today_pct:+.2f}%",
                }
                for p in summary.positions
            ]
        else:
            pos_rows = [{"Symbol": "No open positions", "Name": "", "Quantity": "",
                         "Avg Cost (€)": "", "Current Price (€)": "", "Market Value (€)": "",
                         "Cost Basis (€)": "", "Unrealised P&L (€)": "",
                         "Unrealised P&L %": "", "Today %": ""}]
        pd.DataFrame(pos_rows).to_excel(writer, sheet_name="Positions", index=False)

        # ── Sheet 3: Transactions ────────────────────────────────────────────
        if transactions:
            tx_rows = [
                {
                    "Timestamp": t["timestamp"][:19].replace("T", " "),
                    "Symbol": t["symbol"],
                    "Action": t["action"],
                    "Quantity": round(t["quantity"], 4),
                    "Price USD": round(t["price_usd"], 4),
                    "Price EUR": round(t["price_eur"], 4),
                    "EUR/USD Rate": round(t["eur_usd_rate"], 4),
                    "Total EUR": round(t["total_eur"], 2),
                    "Balance After (€)": round(t["balance_after_eur"], 2),
                    "Notes": t.get("notes") or "",
                }
                for t in transactions
            ]
        else:
            tx_rows = [{"Timestamp": "No transactions yet", "Symbol": "", "Action": "",
                        "Quantity": "", "Price USD": "", "Price EUR": "",
                        "EUR/USD Rate": "", "Total EUR": "", "Balance After (€)": "", "Notes": ""}]
        pd.DataFrame(tx_rows).to_excel(writer, sheet_name="Transactions", index=False)

        # ── Sheet 4: Research ────────────────────────────────────────────────
        if research:
            import json
            res_rows = []
            for r in research:
                try:
                    data = json.loads(r["analysis"])
                except Exception:
                    data = {}
                res_rows.append({
                    "Date": r["created_at"][:10],
                    "Symbol": r["symbol"],
                    "Company": r.get("company_name") or "",
                    "Recommendation": r["recommendation"],
                    "Confidence": r["confidence"],
                    "Target Price USD": r.get("target_price_usd") or "",
                    "Thesis": data.get("thesis") or "",
                    "Bull Case": data.get("bull_case") or "",
                    "Bear Case": data.get("bear_case") or "",
                    "Time Horizon": data.get("time_horizon") or "",
                    "Suggested Position (€)": data.get("suggested_position_size_eur") or "",
                })
        else:
            res_rows = [{"Date": "No research yet", "Symbol": "", "Company": "",
                         "Recommendation": "", "Confidence": "", "Target Price USD": "",
                         "Thesis": "", "Bull Case": "", "Bear Case": "",
                         "Time Horizon": "", "Suggested Position (€)": ""}]
        pd.DataFrame(res_rows).to_excel(writer, sheet_name="Research", index=False)

    # ── Post-process: apply styles via openpyxl ──────────────────────────────
    wb = load_workbook(path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cols = [cell.value for cell in ws[1]]
        _header_style(ws, 1, cols)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            bg = _ALT_ROW if row_idx % 2 == 0 else _WHITE
            fill = PatternFill("solid", fgColor=bg)
            for cell in row:
                cell.fill = fill
                cell.alignment = Alignment(vertical="center")
        _auto_width(ws)
        ws.freeze_panes = "A2"

    wb.save(path)
    return path
